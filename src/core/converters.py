"""Классы ядра для конвертации школьных Excel-журналов.

Модуль реализует исходную бизнес-логику практически без изменений:
определение блоков учеников, обработку объединенных ячеек с датами,
вычисление триместров, извлечение оценок и применение teacher mapping.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.core.models import ConversionStats, TeacherOverride
from src.utils.constants import (
    DEFAULT_HOMEWORK_TEXT,
    DEFAULT_LESSON_COLUMNS,
    DEFAULT_SUBJECT_NAME,
    DEFAULT_TEACHER_NAME,
    DEFAULT_TOPIC_NAME,
    EMPTY_TRIMESTER_MARKERS,
    GRADES_COL_START,
    MONTHS_MAP,
    OUTPUT_GRADES_SHEET_NAME,
    OUTPUT_LESSONS_SHEET_NAME,
)
from src.utils.helpers import (
    extract_class_from_filename,
    extract_date_from_filename,
    generate_lesson_id,
    get_trimester_by_date,
)
from src.utils.patterns import (
    DAY_MONTH_PATTERN,
    GROUP_NUMBER_PATTERN,
    SUBJECT_FALLBACK_PATTERNS,
    SUBJECT_TRAILING_DIGITS_PATTERN,
    TEACHER_LINE_PATTERN,
)

LOGGER = logging.getLogger(__name__)


class DataExtractor:
    """Низкоуровневый помощник для навигации по листу Excel."""

    def __init__(self, worksheet: Worksheet) -> None:
        self.worksheet = worksheet

    def find_cell_coordinates(
        self,
        search_text: str,
        rows_limit: int = 50,
        cols_limit: int = 25,
    ) -> tuple[int | None, int | None]:
        """Находит координаты ячейки, содержащей указанную подстроку."""

        for row_idx in range(1, min(rows_limit + 1, self.worksheet.max_row + 1)):
            for col_idx in range(1, min(cols_limit + 1, self.worksheet.max_column + 1)):
                cell_value = self.worksheet.cell(row=row_idx, column=col_idx).value
                if cell_value and isinstance(cell_value, str) and search_text in cell_value:
                    return row_idx, col_idx
        return None, None

    def get_block_rows_safe(
        self,
        block_start_row: int,
        max_search_rows: int = 100,
    ) -> tuple[int, int]:
        """Определяет последнюю строку блока учеников."""

        end_row = self.worksheet.max_row
        for row_idx in range(
            block_start_row,
            min(block_start_row + max_search_rows, self.worksheet.max_row + 1),
        ):
            name_value = self.worksheet.cell(row=row_idx, column=2).value
            number_value = self.worksheet.cell(row=row_idx, column=1).value

            if not name_value or (isinstance(name_value, str) and not name_value.strip()):
                end_row = row_idx - 1
                break

            if (
                number_value
                and isinstance(number_value, str)
                and ("№" in number_value or "Обучающийся" in number_value)
            ):
                end_row = row_idx - 1
                break

        return block_start_row, end_row


class SchoolJournalConverter:
    """Конвертер одного листа школьного журнала в нормализованные таблицы."""

    def __init__(
        self,
        worksheet: Worksheet,
        sheet_name: str,
        file_name: str | None = None,
        class_from_file: str | None = None,
        academic_year_start: int | None = None,
        overrides: list[TeacherOverride] | None = None,
    ) -> None:
        self.worksheet = worksheet
        self.sheet_name = sheet_name
        self.file_name = file_name
        self.class_from_file = class_from_file
        self.academic_year_start = academic_year_start or datetime.now().year
        self.overrides = overrides or []

        self.extractor = DataExtractor(worksheet)
        self.lessons_registry: dict[str, list[dict[str, Any]]] = {}
        self.students_data: dict[str, dict[str, list[int | str]]] = {}
        self.trimester_grades: dict[str, dict[int, int | None]] = {}

        self.is_valid = True
        self.error_message = ""
        self.lessons_cols: dict[str, int] = dict(DEFAULT_LESSON_COLUMNS)
        self.grades_col_end = 25
        self.blocks: list[dict[str, Any]] = []
        self.class_num = "?"
        self.teacher = DEFAULT_TEACHER_NAME
        self.subject = DEFAULT_SUBJECT_NAME

    def get_year_for_month(self, month: int) -> int:
        """Возвращает календарный год для месяца внутри учебного года."""

        return self.academic_year_start if month >= 9 else self.academic_year_start + 1

    def find_lesson_columns(self) -> None:
        """Ищет реальные колонки даты, темы и домашнего задания."""

        found: dict[str, int | None] = {"дата": None, "тема": None, "дз": None}

        for row_idx in [1, 2, 3]:
            for col_idx in range(15, min(50, self.worksheet.max_column + 1)):
                value = self.worksheet.cell(row=row_idx, column=col_idx).value
                if value and isinstance(value, str):
                    lowered = value.strip().lower()
                    if "дата" in lowered:
                        found["дата"] = col_idx
                    elif "тема" in lowered:
                        found["тема"] = col_idx
                    elif "домашн" in lowered or lowered in ["д/з", "дз"]:
                        found["дз"] = col_idx
            if found["дата"]:
                break

        if not found["дата"]:
            found["дата"] = 21
        if not found["тема"]:
            found["тема"] = found["дата"] + 1
        if not found["дз"]:
            found["дз"] = found["тема"] + 1

        self.lessons_cols = {
            "дата": int(found["дата"]),
            "тема": int(found["тема"]),
            "дз": int(found["дз"]),
        }
        self.grades_col_end = self.lessons_cols["дата"]

    @staticmethod
    def _is_stop_column(value_str: str) -> bool:
        """Определяет, означает ли колонка окончание области оценок."""

        if not value_str:
            return False
        lowered = value_str.strip().lower()
        return lowered == "о" or any(
            marker in lowered
            for marker in [
                "дата",
                "тема",
                "домашн",
                "итого",
                "четверть",
                "полугодие",
                "год",
                "триместр",
            ]
        )

    def extract_subject_fallback(self) -> str:
        """Извлекает предмет из имени листа, если явная ячейка не найдена."""

        sheet = self.sheet_name.strip()
        sheet = SUBJECT_TRAILING_DIGITS_PATTERN.sub("", sheet).strip()
        for pattern in SUBJECT_FALLBACK_PATTERNS:
            match = pattern.match(sheet)
            if match:
                return match.group(1).strip()
        return sheet.split()[0] if sheet.split() else sheet

    def _process_subject_name(self, raw_text: str) -> str:
        """Нормализует предмет с учетом групп и полового деления."""

        if not raw_text:
            return DEFAULT_SUBJECT_NAME

        base_subject = raw_text.split(",")[-1].strip() if "," in raw_text else raw_text.strip()
        raw_text_lower = raw_text.lower()
        base_subject_lower = base_subject.lower()

        if "информатика" in base_subject_lower:
            group_match = GROUP_NUMBER_PATTERN.search(raw_text_lower)
            if group_match:
                base_subject = f"{base_subject} {group_match.group(1)}гр"

        gender_prefix = ""
        if "мальчики" in raw_text_lower:
            gender_prefix = "Мальчики "
        elif "девочки" in raw_text_lower:
            gender_prefix = "Девочки "

        if gender_prefix and any(
            marker in base_subject_lower
            for marker in [
                "труд",
                "технолог",
                "физическая",
                "физ-ра",
                "физкультура",
                "физ.",
            ]
        ):
            if not base_subject_lower.startswith(gender_prefix.lower().strip()):
                base_subject = f"{gender_prefix}{base_subject}"

        return base_subject

    def extract_metadata(self) -> None:
        """Извлекает класс, учителя и предмет, затем применяет overrides."""

        self.class_num = self.class_from_file or "?"

        teacher_row, teacher_col = self.extractor.find_cell_coordinates("Учитель:")
        if teacher_row and teacher_col:
            teacher_cell_value = self.worksheet.cell(row=teacher_row, column=teacher_col).value
            teacher_match = TEACHER_LINE_PATTERN.search(str(teacher_cell_value))
            if teacher_match:
                self.teacher = teacher_match.group(1).strip()
            else:
                self.teacher = str(teacher_cell_value).replace("Учитель:", "").strip()

            raw_subject_value = (
                self.worksheet.cell(row=teacher_row - 2, column=teacher_col).value
                if teacher_row > 2
                else None
            )
            self.subject = (
                self._process_subject_name(str(raw_subject_value))
                if raw_subject_value
                else self.extract_subject_fallback()
            )
        else:
            self.teacher = DEFAULT_TEACHER_NAME
            self.subject = self.extract_subject_fallback()

        if self.overrides:
            current_class_lower = str(self.class_num).strip().lower()
            current_subject_lower = str(self.subject).strip().lower()

            for override in self.overrides:
                override_class = override.class_name.strip().lower()
                override_subject_keyword = override.subject.strip().lower()
                if (
                    override_class == current_class_lower
                    and override_subject_keyword
                    and override_subject_keyword in current_subject_lower
                ):
                    self.teacher = override.teacher.strip()
                    break

    def collect_all_lessons(self) -> None:
        """Собирает реестр уроков с полными датами, темами и ДЗ."""

        self.find_lesson_columns()

        for row_idx in range(1, self.worksheet.max_row + 1):
            date_cell = self.worksheet.cell(row=row_idx, column=self.lessons_cols["дата"]).value
            if not date_cell or not isinstance(date_cell, str):
                continue

            date_str = date_cell.strip()
            if "учитель" in date_str.lower() or "." not in date_str:
                continue

            try:
                parts = date_str.split(".")
                if len(parts) >= 2:
                    day_str, month_str = parts[0], parts[1]
                    if day_str.isdigit() and month_str.isdigit():
                        day = int(day_str)
                        month = int(month_str)
                        if 1 <= day <= 31 and 1 <= month <= 12:
                            year = self.get_year_for_month(month)
                            full_date_obj = datetime(year, month, day)
                            full_date_str = full_date_obj.strftime("%Y-%m-%d")
                            lesson_trimester = get_trimester_by_date(full_date_obj)

                            topic_value = str(
                                self.worksheet.cell(
                                    row=row_idx,
                                    column=self.lessons_cols["тема"],
                                ).value
                                or DEFAULT_TOPIC_NAME
                            ).strip()
                            homework_value = str(
                                self.worksheet.cell(
                                    row=row_idx,
                                    column=self.lessons_cols["дз"],
                                ).value
                                or DEFAULT_HOMEWORK_TEXT
                            ).strip()
                            topic_value = topic_value.replace("\n", " ").replace("\r", " ")
                            homework_value = homework_value.replace("\n", " ").replace("\r", " ")

                            if full_date_str not in self.lessons_registry:
                                self.lessons_registry[full_date_str] = []

                            occurrence_idx = len(self.lessons_registry[full_date_str])
                            lesson_id = generate_lesson_id(
                                self.subject,
                                self.class_num,
                                f"{full_date_str}_{occurrence_idx}",
                            )

                            self.lessons_registry[full_date_str].append(
                                {
                                    "id": lesson_id,
                                    "дата_полная": full_date_str,
                                    "тема": topic_value,
                                    "дз": homework_value,
                                    "триместр": lesson_trimester,
                                }
                            )
            except (ValueError, TypeError):
                continue

    def find_blocks(self) -> None:
        """Находит блоки учеников и сопоставляет колонкам реальные даты."""

        self.blocks = []

        for row_idx in range(1, self.worksheet.max_row + 1):
            cell_a = self.worksheet.cell(row=row_idx, column=1).value
            if cell_a and isinstance(cell_a, str) and cell_a.strip() == "№":
                self.blocks.append(
                    {
                        "start_row": row_idx,
                        "header_row": row_idx + 1,
                        "students_start_row": row_idx + 2,
                    }
                )

        if not self.blocks:
            self.is_valid = False
            self.error_message = "Блоки учеников не найдены"
            return

        for block in self.blocks:
            _, end_row = self.extractor.get_block_rows_safe(block["students_start_row"])
            block["end_row"] = end_row
            header_row = block["header_row"]
            month_months: dict[int, list[dict[str, Any]]] = {}
            current_month: int | None = None

            merged_headers: dict[int, int] = {}
            for merged_range in self.worksheet.merged_cells.ranges:
                if (
                    merged_range.min_row <= header_row <= merged_range.max_row
                    and merged_range.min_col < merged_range.max_col
                ):
                    for col_idx in range(merged_range.min_col + 1, merged_range.max_col + 1):
                        merged_headers[col_idx] = merged_range.min_col

            last_valid_day: int | None = None

            for col_idx in range(
                GRADES_COL_START,
                min(self.grades_col_end, self.worksheet.max_column + 1),
            ):
                value_above = self.worksheet.cell(row=header_row - 1, column=col_idx).value
                value_current = self.worksheet.cell(row=header_row, column=col_idx).value

                should_break = False

                if value_above and isinstance(value_above, str):
                    value_str = value_above.strip().lower()
                    if value_str in MONTHS_MAP:
                        current_month = MONTHS_MAP[value_str]
                    elif value_str == "а":
                        current_month = 4
                    elif value_str == "и":
                        current_month = 6
                    elif value_str == "м":
                        if current_month in [1, 2]:
                            current_month = 3
                        elif current_month in [4, 5]:
                            current_month = 5

                    if self._is_stop_column(value_str):
                        should_break = True

                if value_current and isinstance(value_current, str) and self._is_stop_column(value_current):
                    should_break = True

                if should_break:
                    break

                if current_month:
                    day_value = value_current
                    try:
                        if day_value is not None:
                            day_str = str(day_value).strip()
                            if day_str.isdigit():
                                day = int(day_str)
                                if 1 <= day <= 31:
                                    last_valid_day = day
                                    month_months.setdefault(current_month, []).append(
                                        {
                                            "day": day,
                                            "col": col_idx,
                                            "col_letter": get_column_letter(col_idx),
                                        }
                                    )
                            else:
                                last_valid_day = None
                        else:
                            if col_idx in merged_headers:
                                parent_col = merged_headers[col_idx]
                                parent_value = self.worksheet.cell(
                                    row=header_row,
                                    column=parent_col,
                                ).value
                                if (
                                    parent_value is not None
                                    and str(parent_value).strip().isdigit()
                                ):
                                    day = int(str(parent_value).strip())
                                    if 1 <= day <= 31:
                                        month_months.setdefault(current_month, []).append(
                                            {
                                                "day": day,
                                                "col": col_idx,
                                                "col_letter": get_column_letter(col_idx),
                                            }
                                        )
                            elif last_valid_day is not None:
                                has_grades = False
                                search_limit = min(end_row + 1, self.worksheet.max_row + 1)
                                for row_idx in range(block["students_start_row"], search_limit):
                                    cell_value = self.worksheet.cell(row=row_idx, column=col_idx).value
                                    if cell_value and str(cell_value).strip().lower() in [
                                        "2",
                                        "3",
                                        "4",
                                        "5",
                                        "н",
                                    ]:
                                        has_grades = True
                                        break
                                if has_grades:
                                    month_months.setdefault(current_month, []).append(
                                        {
                                            "day": last_valid_day,
                                            "col": col_idx,
                                            "col_letter": get_column_letter(col_idx),
                                        }
                                    )
                    except (ValueError, TypeError):
                        continue

            block["months"] = month_months

    def process_students_across_blocks(self) -> None:
        """Распределяет оценки учеников по найденным урокам."""

        for block in self.blocks:
            students_start_row = int(block["students_start_row"])
            end_row = int(block["end_row"])

            block_students = {
                row_idx: name.strip()
                for row_idx in range(students_start_row, end_row + 1)
                if (
                    name := self.worksheet.cell(row=row_idx, column=2).value
                ) and isinstance(name, str) and name.strip()
            }

            date_counts: dict[str, int] = {}
            col_to_lesson_id: dict[int, str] = {}

            for month_num, days_list in block["months"].items():
                for day_info in sorted(days_list, key=lambda item: item["col"]):
                    year = self.get_year_for_month(month_num)
                    full_date_str = f"{year}-{month_num:02d}-{day_info['day']:02d}"

                    if full_date_str not in date_counts:
                        date_counts[full_date_str] = 0
                    lesson_index = date_counts[full_date_str]
                    date_counts[full_date_str] += 1

                    if full_date_str in self.lessons_registry:
                        lessons_for_date = self.lessons_registry[full_date_str]
                        lesson = lessons_for_date[min(lesson_index, len(lessons_for_date) - 1)]
                        col_to_lesson_id[int(day_info["col"])] = str(lesson["id"])

            for student_row, student_name in block_students.items():
                if student_name not in self.students_data:
                    self.students_data[student_name] = {}

                for _, days_list in block["months"].items():
                    for day_info in days_list:
                        col_idx = int(day_info["col"])
                        if col_idx not in col_to_lesson_id:
                            continue

                        lesson_id = col_to_lesson_id[col_idx]
                        cell_value = self.worksheet.cell(row=student_row, column=col_idx).value

                        if cell_value is not None:
                            value_str = str(cell_value).strip().lower()
                            if not value_str:
                                continue

                            grades_found = DAY_MONTH_PATTERN.findall(value_str)
                            for grade_str in grades_found:
                                grade_value: int | str = "н" if grade_str.lower() == "н" else int(grade_str)
                                self.students_data[student_name].setdefault(lesson_id, []).append(
                                    grade_value,
                                )

    def extract_trimester_grades(self) -> None:
        """Извлекает итоговые оценки по триместрам."""

        for block in self.blocks:
            students_start_row = int(block["students_start_row"])
            end_row = int(block["end_row"])
            header_row = int(block["header_row"])

            current_month: int | None = None

            for col_idx in range(
                GRADES_COL_START,
                min(self.grades_col_end + 10, self.worksheet.max_column + 1),
            ):
                col_header = self.worksheet.cell(row=header_row, column=col_idx).value
                value_above = self.worksheet.cell(row=header_row - 1, column=col_idx).value

                if value_above and isinstance(value_above, str):
                    month_str = value_above.strip().lower()
                    if month_str in MONTHS_MAP:
                        current_month = MONTHS_MAP[month_str]
                    elif month_str == "а":
                        current_month = 4
                    elif month_str == "и":
                        current_month = 6
                    elif month_str == "н":
                        current_month = 11
                    elif month_str == "м":
                        if current_month in [1, 2, 3, 4]:
                            current_month = 5
                        else:
                            current_month = 3

                trimester_num: int | None = None

                if col_header and isinstance(col_header, str):
                    header_str = col_header.strip().lower()

                    if header_str in ["вт", "чт", "пт", "ср", "сб", "вс", "от"]:
                        continue

                    if (
                        "т" in header_str
                        or "итог" in header_str
                        or "тр" in header_str
                        or "п" in header_str
                    ):
                        if "1" in header_str:
                            trimester_num = 1
                        elif "2" in header_str:
                            trimester_num = 2
                        elif "3" in header_str:
                            trimester_num = 3
                        elif current_month is not None:
                            if current_month in [9, 10, 11]:
                                trimester_num = 1
                            elif current_month in [12, 1, 2, 3, 4]:
                                trimester_num = 2
                            elif current_month in [5, 6, 7, 8]:
                                trimester_num = 3

                if trimester_num is not None:
                    for row_idx in range(students_start_row, end_row + 1):
                        name_cell = self.worksheet.cell(row=row_idx, column=2).value
                        grade_cell = self.worksheet.cell(row=row_idx, column=col_idx).value

                        if not (name_cell and isinstance(name_cell, str)):
                            continue

                        name = name_cell.strip()
                        if not name:
                            continue

                        self.trimester_grades.setdefault(name, {})

                        if grade_cell is None:
                            continue

                        grade_str = str(grade_cell).strip().lower()
                        if not grade_str:
                            continue

                        if grade_str in EMPTY_TRIMESTER_MARKERS:
                            self.trimester_grades[name][trimester_num] = None
                            continue

                        try:
                            grade = int(grade_cell)
                            if 2 <= grade <= 5:
                                self.trimester_grades[name][trimester_num] = grade
                        except (ValueError, TypeError):
                            continue

    def convert(self) -> bool:
        """Запускает полный цикл обработки листа."""

        self.extract_metadata()
        self.collect_all_lessons()
        self.find_blocks()
        if not self.is_valid:
            return False
        self.process_students_across_blocks()
        self.extract_trimester_grades()
        return bool(self.students_data)

    def get_dataframes(self) -> tuple[pd.DataFrame, pd.DataFrame]:
        """Возвращает итоговые DataFrame по оценкам и урокам."""

        flat_records: list[dict[str, Any]] = []
        flat_lessons: list[dict[str, Any]] = []

        for lessons_list in self.lessons_registry.values():
            flat_lessons.extend(lessons_list)

        lesson_lookup = {lesson["id"]: lesson for lesson in flat_lessons}

        for student_name, lessons_grades in self.students_data.items():
            trimester_grades_dict = self.trimester_grades.get(student_name, {})

            for lesson_id, grades_list in lessons_grades.items():
                lesson_details = lesson_lookup.get(lesson_id)
                if lesson_details:
                    lesson_trimester = lesson_details["триместр"]
                    trimester_grade = trimester_grades_dict.get(lesson_trimester, None)

                    for grade in grades_list:
                        flat_records.append(
                            {
                                "ФИО": student_name,
                                "Класс": self.class_num,
                                "Предмет": self.subject,
                                "Учитель": self.teacher,
                                "Оценка": grade,
                                "ID урока": lesson_id,
                                "Оценка за триместр": trimester_grade,
                                "Номер триместра": lesson_trimester,
                            }
                        )

        df_records = pd.DataFrame(flat_records)
        if not df_records.empty:
            df_records = df_records.sort_values(["Класс", "Предмет", "ФИО", "ID урока"])

        df_lessons = pd.DataFrame(flat_lessons)
        if not df_lessons.empty:
            df_lessons["Класс"] = self.class_num
            df_lessons["Предмет"] = self.subject
            final_columns = ["id", "Класс", "Предмет", "дата_полная", "тема", "дз"]
            df_lessons = df_lessons[final_columns].rename(
                columns={
                    "id": "ID урока",
                    "дата_полная": "Дата урока",
                    "тема": "Тема урока",
                    "дз": "Домашнее задание",
                },
            )
            df_lessons = df_lessons.sort_values(["Класс", "Предмет", "Дата урока"])

        return df_records, df_lessons


class SchoolJournalFileConverter:
    """Обрабатывает все листы одного Excel-файла журнала."""

    def __init__(
        self,
        excel_file: str | Path,
        overrides: list[TeacherOverride] | None = None,
    ) -> None:
        self.file_path = Path(excel_file)
        self.file_name = self.file_path.name
        self.workbook = openpyxl.load_workbook(str(excel_file), data_only=True)
        self.sheet_results: dict[str, dict[str, Any]] = {}
        self.all_records: list[pd.DataFrame] = []
        self.all_lessons: list[pd.DataFrame] = []
        self.class_from_file = extract_class_from_filename(self.file_name)
        file_date = extract_date_from_filename(self.file_name)
        self.academic_year_start = file_date.year - 1 if file_date.month < 9 else file_date.year
        self.overrides = overrides or []

    def convert_all_sheets(self) -> bool:
        """Конвертирует все листы рабочей книги."""

        for sheet_name in self.workbook.sheetnames:
            try:
                converter = SchoolJournalConverter(
                    self.workbook[sheet_name],
                    sheet_name,
                    file_name=self.file_name,
                    class_from_file=self.class_from_file,
                    academic_year_start=self.academic_year_start,
                    overrides=self.overrides,
                )

                if converter.convert():
                    df_records, df_lessons = converter.get_dataframes()
                    self.sheet_results[sheet_name] = {
                        "status": "успешно",
                        "records_count": len(df_records),
                    }
                    self.all_records.append(df_records)
                    if not df_lessons.empty:
                        self.all_lessons.append(df_lessons)
                elif not converter.is_valid:
                    self.sheet_results[sheet_name] = {
                        "status": "пропущено",
                        "reason": converter.error_message,
                        "records_count": 0,
                    }
                else:
                    self.sheet_results[sheet_name] = {
                        "status": "пропущено",
                        "reason": "данных не найдено",
                        "records_count": 0,
                    }
            except Exception as error:  # noqa: BLE001
                LOGGER.exception(
                    "Ошибка обработки листа '%s' файла '%s': %s",
                    sheet_name,
                    self.file_name,
                    error,
                )
                self.sheet_results[sheet_name] = {
                    "status": "ошибка",
                    "error": str(error),
                    "records_count": 0,
                }

        return len(self.all_records) > 0

    def get_combined_dataframes(self) -> tuple[pd.DataFrame, pd.DataFrame]:
        """Объединяет результаты всех листов файла."""

        df_combined = pd.concat(self.all_records, ignore_index=True) if self.all_records else pd.DataFrame()

        df_lessons = pd.DataFrame()
        if self.all_lessons:
            df_lessons = pd.concat(self.all_lessons, ignore_index=True)
            df_lessons = df_lessons.drop_duplicates(
                subset=["ID урока"],
                keep="first",
            ).reset_index(drop=True)

        return df_combined, df_lessons

    def get_stats(self) -> ConversionStats:
        """Возвращает статистику обработки файла."""

        return ConversionStats(
            file_name=self.file_name,
            total_sheets=len(self.workbook.sheetnames),
            processed_sheets=sum(
                1
                for sheet_data in self.sheet_results.values()
                if sheet_data["status"] == "успешно"
            ),
            records_count=sum(
                sheet_data.get("records_count", 0)
                for sheet_data in self.sheet_results.values()
            ),
        )


class BatchFolderConverter:
    """Пакетный конвертер Excel-файлов из одной папки."""

    def __init__(
        self,
        input_folder: str | Path,
        overrides: list[TeacherOverride] | None = None,
    ) -> None:
        self.input_folder = Path(input_folder)
        if not self.input_folder.is_dir():
            raise ValueError(f"Папка не найдена: {input_folder}")

        self.excel_files = sorted(self.input_folder.glob("*.xlsx"))
        if not self.excel_files:
            raise ValueError(f"Excel файлы не найдены в папке: {input_folder}")

        self.file_results: dict[str, dict[str, Any]] = {}
        self.all_records: list[pd.DataFrame] = []
        self.all_lessons: list[pd.DataFrame] = []
        self.overrides = overrides or []

    def convert_all_files(self) -> bool:
        """Конвертирует все Excel-файлы в папке."""

        LOGGER.info("=" * 100)
        LOGGER.info("ПАКЕТНОЕ ПРЕОБРАЗОВАНИЕ")
        LOGGER.info("=" * 100)

        for file_index, excel_file in enumerate(self.excel_files, start=1):
            LOGGER.info(
                "[%s/%s] Обработка файла: '%s'",
                file_index,
                len(self.excel_files),
                excel_file.name,
            )
            try:
                converter = SchoolJournalFileConverter(excel_file, overrides=self.overrides)
                if converter.convert_all_sheets():
                    df_file, df_lessons_file = converter.get_combined_dataframes()
                    self.all_records.append(df_file)
                    if not df_lessons_file.empty:
                        self.all_lessons.append(df_lessons_file)

                    stats = converter.get_stats()
                    self.file_results[excel_file.name] = {
                        "status": "успешно",
                        "stats": stats.to_dict(),
                    }
                    LOGGER.info("  Успешно: %s записей", stats.records_count)
                else:
                    self.file_results[excel_file.name] = {"status": "пропущено"}
                    LOGGER.warning("  Пропущено")
            except Exception as error:  # noqa: BLE001
                LOGGER.exception("Ошибка обработки файла '%s': %s", excel_file.name, error)
                self.file_results[excel_file.name] = {
                    "status": "ошибка",
                    "error": str(error),
                }

        return len(self.all_records) > 0

    def save_results(self, output_file: str | Path) -> bool:
        """Сохраняет объединенный результат в Excel."""

        if not self.all_records:
            return False

        try:
            df_combined = pd.concat(self.all_records, ignore_index=True)
            df_combined = df_combined.sort_values(["Класс", "Предмет", "ФИО", "ID урока"])

            df_lessons_combined = pd.DataFrame()
            if self.all_lessons:
                df_lessons_combined = pd.concat(self.all_lessons, ignore_index=True)
                df_lessons_combined = (
                    df_lessons_combined.drop_duplicates(
                        subset=["ID урока"],
                        keep="first",
                    )
                    .reset_index(drop=True)
                    .sort_values(["Класс", "Предмет", "Дата урока"])
                )
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                df_combined.to_excel(
                    writer,
                    sheet_name=OUTPUT_GRADES_SHEET_NAME,
                    index=False,
                )
                if not df_lessons_combined.empty:
                    df_lessons_combined.to_excel(
                        writer,
                        sheet_name=OUTPUT_LESSONS_SHEET_NAME,
                        index=False,
                    )

            LOGGER.info("Сохранено: %s", output_file)
            return True
        except Exception as error:  # noqa: BLE001
            LOGGER.exception("Ошибка сохранения результата: %s", error)
            return False
